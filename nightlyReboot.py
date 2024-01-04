import os
from dotenv import load_dotenv
import time
import datetime
import sched
import requests
from urllib3.exceptions import InsecureRequestWarning
from openpyxl import load_workbook
import concurrent.futures

#Loading environment variables
load_dotenv()
PASSWORD = os.environ.get("PASSWORD")
FILENAME = os.environ.get("FILENAME")

# disable ssl warning
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

#Loading excel workbook with IPs
codecList = load_workbook(FILENAME)
codecSheet = codecList.active


#Defining reboot function
DEFAULT_PASSWORD = 'Basic ' + PASSWORD

def reboot_request(auth_val=DEFAULT_PASSWORD, ip=None, idx=None):
    url = f"https://{ip}/putxml"
    cell_no = idx+2
    payload = "<Command>\r\n\t<Standby>\r\n\t\t<Deactivate></Deactivate>\r\n\t</Standby>\r\n\t<UserInterface>\r\n\t\t<Message>\r\n\t\t\t<Alert>\r\n\t\t\t\t<Display>\r\n\t\t\t\t\t<Duration>10</Duration>\r\n\t\t\t\t\t<Text>Use touchpanel to cancel reboot</Text>\r\n\t\t\t\t\t<Title>NIGHTLY REBOOT INITIATED</Title>\r\n\t\t\t\t</Display>\r\n\t\t\t</Alert>\r\n\t\t</Message>\r\n\t</UserInterface>\r\n</Command>"

    headers = {'Content-Type': 'text/xml','Authorization': auth_val}
    try:
        response = requests.request("POST", url, headers=headers, data=payload, verify=False)
        response.raise_for_status()
        print(response.status_code)
        codecSheet[f"D{cell_no}"] = response.status_code
    except requests.exceptions.HTTPError as err:
        print(err.response.status_code)
        codecSheet[f"D{cell_no}"] = err.response.status_code

def initiate_reboot():
    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
        for idx, value in enumerate(codecSheet.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True)):
            ip = value[0]
            executor.submit(reboot_request, DEFAULT_PASSWORD, ip, idx)
    codecList.save("output.xlsx")


#Setting up reboot timer
s = sched.scheduler(time.time, time.sleep)

def nightly_reboot():
    now = datetime.datetime.now()
    day = datetime.date(now.year, now.month, now.day).isoweekday()
    hour = now.hour
    min = now.minute
    sec = now.second
    #Interval for when the script checks the time in seconds. Make sure the interval is guaranteed to check during reboot trigger window
    # if day == 6:
    interval = 10
    # else:
    #     interval = 11 * 60 * 60

    print(f'{hour} {min} {sec}')
    if hour == 2 and min == 00 and sec >= 30: #Time window for reboot trigger
        print('Reboot initiated')
        initiate_reboot()
        time.sleep(60)  #Delay for restarting timer. Make sure it is enough time to exit reboot trigger window
        s.enter(interval, 1, nightly_reboot, ())
    else:
        s.enter(interval, 1, nightly_reboot, ()) 

if __name__ == "__main__":
    s.enter(0, 1, nightly_reboot, ())
    s.run()

