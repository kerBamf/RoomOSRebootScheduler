import os
from dotenv import load_dotenv
import time
import datetime
import sched
import requests
from urllib3.exceptions import InsecureRequestWarning
from openpyxl import load_workbook
import openpyxl
import concurrent.futures
import smtplib
from email.message import EmailMessage

#Loading environment variables
load_dotenv()
PASSWORD = os.environ.get("PASSWORD")
FILENAME = os.environ.get("FILENAME")

# disable ssl warning
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

#Building email function for output report

def send_email():

    text = 'Good Morning,\n Here is the report from the previous automated reboot.'
    msg = EmailMessage()
    msg['From'] = 'codec_reboot_report_noreply@mskcc.org'
    msg['To'] = 'pedigoz@mskcc.org'
    msg['Subject'] = 'Codec Reboot Report'
    msg.set_content(text)
    server = 'exchange2007.mskcc.org'
    port = 25
    excel_file = 'output.xlsx'
    file_data = open(excel_file, 'rb').read()
    msg.add_attachment(file_data, maintype='application', subtype='xlsx', filename=excel_file)
    smtp = smtplib.SMTP(server, port)
    smtp.send_message(msg)
    smtp.quit()

#Defining reboot function
DEFAULT_PASSWORD = 'Basic ' + PASSWORD

def reboot_request(auth_val=DEFAULT_PASSWORD, ip=None, idx=None, file=None):
    url = f"https://{ip}/putxml"
    cell_no = idx+2
    payload = "<Command>\r\n\t<Standby>\r\n\t\t<Deactivate></Deactivate>\r\n\t</Standby>\r\n\t<UserInterface>\r\n\t\t<Message>\r\n\t\t\t<Alert>\r\n\t\t\t\t<Display>\r\n\t\t\t\t\t<Duration>10</Duration>\r\n\t\t\t\t\t<Text>Use touchpanel to cancel reboot</Text>\r\n\t\t\t\t\t<Title>AUTOMATED REBOOT INITIATED</Title>\r\n\t\t\t\t</Display>\r\n\t\t\t</Alert>\r\n\t\t</Message>\r\n\t</UserInterface>\r\n</Command>"

    headers = {'Content-Type': 'text/xml','Authorization': auth_val}
    try:
        response = requests.request("POST", url, headers=headers, data=payload, verify=False)
        response.raise_for_status()
        print(response.status_code)
        file[f"D{cell_no}"] = response.status_code
    except requests.exceptions.HTTPError as err:
        print(err.response.status_code)
        file[f"D{cell_no}"] = err.response.status_code

def initiate_reboot(excel_file):
    try:
        #Loading excel workbook with IPs
        codecList = load_workbook(excel_file)
        codecSheet = codecList.active
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            for idx, value in enumerate(codecSheet.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True)):
                ip = value[0]
                executor.submit(reboot_request, DEFAULT_PASSWORD, ip, idx, codecSheet)
        codecList.save("output.xlsx")
        #Consider adding ping function for later version here
        send_email()
    except Exception as error:
        print(error)



#Setting up reboot timer
s = sched.scheduler(time.time, time.sleep)

def weekly_reboot():
    now = datetime.datetime.now()
    day = datetime.date(now.year, now.month, now.day).isoweekday()
    hour = now.hour
    min = now.minute
    sec = now.second
    interval = None
    
    #Interval for when the script checks the time in seconds. Make sure the interval is guaranteed to check during reboot trigger window
    if day == 6:
        interval = 10
    else:
        interval = 60 * 60

    print(f'{hour} {min} {sec}')
    if day == 6 and hour == 2 and min == 30 and sec >= 30: #Time window for reboot trigger
        print('Reboot initiated')
        initiate_reboot(FILENAME)
        time.sleep(60)  #Delay for restarting timer. Make sure it is enough time to exit reboot trigger window
        s.enter(interval, 1, weekly_reboot, ())
    else:
        s.enter(interval, 1, weekly_reboot, ()) 

if __name__ == "__main__":
    s.enter(0, 1, weekly_reboot, ())
    s.run()

